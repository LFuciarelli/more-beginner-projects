from openpyxl import Workbook, load_workbook

# Creates a new workbook named "excel module" only if it doesn't already exist

try:
    wb = load_workbook("excel_module.xlsx")
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "expense_tracker"
    sheet["A1"] = "Expense"
    sheet["B1"] = "Cost"
    sheet["C1"] = "Date"

while True:
    expense = input("What was the expense? [E.g. T-shirt]")
    cost = input("How much did it cost? [E.g. 10.00]")
    date = input("What day was it on? [dd/mm/yy]")
    row = sheet.max_row + 1         # The first empty row
    sheet[f"A{row}"] = expense
    sheet[f"B{row}"] = cost
    sheet[f"C{row}"] = date
    asw = input("Do you want to continue? [y/n]").lower().strip()
    if asw == "n":
        break

wb.save(filename="excel_module.xlsx")
